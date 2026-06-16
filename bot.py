
import os
import re
import json
from pathlib import Path
from datetime import datetime
from difflib import SequenceMatcher

import pdfplumber
import gspread
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")
TEMPLATE_FILE = os.getenv("TEMPLATE_FILE", "KP_Client_FireProtect.xlsx")
WORK_DIR = Path(os.getenv("WORK_DIR", "work"))
WORK_DIR.mkdir(exist_ok=True)

GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID")
GOOGLE_KEY_JSON = os.getenv("GOOGLE_KEY_JSON")

BASE_SHEET = "Baza_preturi"
SYN_SHEET = "Sinonime"

AUTO_MATCH_SCORE = 0.78
REVIEW_MATCH_SCORE = 0.58

DEFAULT_WORK_CODES = [
    "LUCR-MONT-TEV", "LUCR-SUD-TEV", "LUCR-MONT-STP", "LUCR-SUD-STP",
    "LUCR-MONT-AUT", "LUCR-PUN-HID", "LUCR-PUN-EL", "DOC-EXEC",
    "LUCR-EXPL", "MATERIALE", "CAZARE", "TRANSPORT",
]

def normalize_text(text: str) -> str:
    if not text:
        return ""
    t = str(text).lower()
    repl = {
        "ă": "a", "â": "a", "î": "i", "ș": "s", "ş": "s", "ț": "t", "ţ": "t",
        "×": "x", "*": "x", "х": "x", "ø": "d", "Ø": "d",
        "оцинкованная": "zincata", "оцинк": "zincata", "zincată": "zincata", "zincat": "zincata",
        "электросварная": "negru", "черная": "negru", "чёрная": "negru",
        "otel negru": "negru", "oțel negru": "negru",
        "țeavă": "teava", "țeava": "teava", "труба": "teava",
        "отвод": "cot", "угол": "cot", "тройник": "teu", "заглушка": "dop", "муфта": "mufa",
        "кран": "robinet", "задвижка": "vana", "затвор": "vana", "манометр": "manometru",
        "спринклер": "sprinkler", "ороситель": "sprinkler", "ду": "dn",
        "consolă": "consola", "brățară": "bratara", "ancoră": "ancora",
        "șaibă": "saiba", "şurub": "surub",
    }
    for a, b in repl.items():
        t = t.replace(a, b)
    t = re.sub(r"[^a-zа-я0-9/.\- x]", " ", t, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", t).strip()

def extract_dn(text: str):
    t = normalize_text(text)
    m = re.search(r"\bdn\s*(\d{2,3})\b", t)
    if m:
        return int(m.group(1))
    m = re.search(r"\b(?:d)?(\d{2,3})\s*x\s*\d", t)
    if m:
        dia = int(m.group(1))
        mp = {21: 15, 26: 20, 27: 20, 32: 25, 33: 25, 42: 32, 48: 40, 57: 50, 60: 50,
              76: 65, 89: 80, 108: 100, 114: 100, 133: 125, 139: 125, 159: 150, 168: 150, 219: 200}
        if dia in mp:
            return mp[dia]
        nearest = min(mp.keys(), key=lambda x: abs(x - dia))
        if abs(nearest - dia) <= 3:
            return mp[nearest]
    return None

def detect_material(text: str):
    t = normalize_text(text)
    if "zincata" in t:
        return "zincata"
    if "negru" in t or "du" in t:
        return "negru"
    return ""

def similarity(a, b):
    a = normalize_text(a)
    b = normalize_text(b)
    if not a or not b:
        return 0.0
    r = SequenceMatcher(None, a, b).ratio()
    ta, tb = set(a.split()), set(b.split())
    overlap = len(ta & tb) / max(1, len(ta | tb))
    return max(r, 0.65 * r + 0.35 * overlap)

def load_database(path):
    wb = load_workbook(path, data_only=False)
    base = wb[BASE_SHEET]
    syn = wb[SYN_SHEET]
    products = {}
    search = []
    for row in base.iter_rows(min_row=2, values_only=True):
        code = row[0]
        if not code:
            continue
        item = {
            "code": str(code).strip(),
            "category": str(row[1] or "").strip(),
            "name": str(row[2] or "").strip(),
            "material": str(row[3] or "").strip(),
            "dn": str(row[4] or "").strip(),
            "um": str(row[5] or "").strip(),
        }
        products[item["code"]] = item
        item["search_text"] = normalize_text(" ".join([
            item["code"], item["category"], item["name"], item["material"],
            "DN" + item["dn"] if item["dn"] else ""
        ]))
        search.append(item)
    synonyms = {}
    for row in syn.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            synonyms[normalize_text(row[0])] = str(row[1]).strip()
    return products, search, synonyms

def extract_pdf_text(pdf_path):
    out = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for p in pdf.pages:
            out.append(p.extract_text() or "")
    return "\n".join(out)

def guess_qty(line):
    pats = [
        r"(\d+(?:[.,]\d+)?)\s*(buc|buc\.|m|ml|m\.l\.|set|шт|м)\b",
        r"\b(buc|buc\.|m|ml|m\.l\.|set|шт|м)\s*(\d+(?:[.,]\d+)?)",
    ]
    for p in pats:
        m = re.search(p, line, flags=re.I)
        if m:
            nums = [g for g in m.groups() if re.match(r"^\d", str(g))]
            if nums:
                return float(nums[0].replace(",", "."))
    return 1

def is_candidate(line):
    n = normalize_text(line)
    ignore_phrases = [
        "reteaua de sprinklere", "retea de sprinklere", "retele de sprinklere",
        "specificatii materiale", "specificatii automatizare", "sistem stingere",
        "total materiale", "total", "nota", "nr poz",
    ]
    if any(p in n for p in ignore_phrases):
        return False
    keys = [
        "teava", "dn", "zincata", "negru", "cot", "teu", "dop", "red",
        "sprinkler", "mufa", "vana", "robinet", "manometru", "flansa",
        "pompa", "rezervor", "cablu", "tija", "piulita", "saiba",
        "bratara", "ancora", "consola",
    ]
    return any(k in n for k in keys)

def find_by_syn(line, synonyms):
    n = normalize_text(line)
    for s, code in sorted(synonyms.items(), key=lambda x: len(x[0]), reverse=True):
        if s and s in n:
            return code, 1.0, "Sinonime"
    return None, 0, ""

def find_best(line, search):
    n = normalize_text(line)
    dn = extract_dn(line)
    mat = detect_material(line)
    best, score_best = None, 0
    for item in search:
        score = similarity(n, item["search_text"])
        item_dn = item.get("dn", "")
        if dn and str(dn) == str(item_dn):
            score += 0.12
        elif dn and item_dn and str(dn) != str(item_dn):
            score -= 0.25
        imat = normalize_text(item.get("material", ""))
        if mat and mat in imat:
            score += 0.10
        elif mat and imat in ["zincata", "negru"] and mat != imat:
            score -= 0.18
        if score > score_best:
            score_best, best = score, item
    if best:
        return best["code"], min(score_best, 1.0), "Baza_preturi fuzzy"
    return None, 0, ""

def recognize(text, products, search, synonyms):
    found, review, unknown = [], [], []
    for line in [x.strip() for x in text.splitlines() if x.strip()]:
        if not is_candidate(line):
            continue
        code, score, method = find_by_syn(line, synonyms)
        if not code:
            code, score, method = find_best(line, search)
        qty = guess_qty(line)
        if code and score >= AUTO_MATCH_SCORE:
            product = products.get(code, {})
            found.append(["OK", code, product.get("name", ""), qty, product.get("um", ""), "", line])
        elif code and score >= REVIEW_MATCH_SCORE:
            product = products.get(code, {})
            review.append(["VERIFICA", code, product.get("name", ""), qty, product.get("um", ""), "", line])
        else:
            unknown.append(["NEIDENTIFICAT", "", "", "", "", "", line])
    merged = {}
    for row in found:
        code = row[1]
        if code not in merged:
            merged[code] = row.copy()
        else:
            merged[code][3] += row[3]
            merged[code][6] += " | " + row[6][:80]
    return list(merged.values()), review, unknown

def add_default_works(rows, products):
    existing = {r[1] for r in rows}
    for code in DEFAULT_WORK_CODES:
        if code in products and code not in existing:
            p = products[code]
            rows.append(["WORK", code, p.get("name", ""), 1, "set", "Lucrari Servicii", "auto lucrari"])
    return rows

def get_gspread_client():
    if not GOOGLE_KEY_JSON:
        raise RuntimeError("Nu este setata variabila GOOGLE_KEY_JSON in Railway.")
    if not GOOGLE_SHEET_ID:
        raise RuntimeError("Nu este setata variabila GOOGLE_SHEET_ID in Railway.")
    info = json.loads(GOOGLE_KEY_JSON)
    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_info(info, scopes=scopes)
    return gspread.authorize(creds)

def write_import_to_google_sheet(rows):
    gc = get_gspread_client()
    sh = gc.open_by_key(GOOGLE_SHEET_ID)
    ws = sh.worksheet("Import")
    ws.clear()
    values = [["Status", "Cod", "Denumire", "Cantitate", "U.M.", "Sectiune", "Text PDF"]]
    values.extend(rows)
    ws.update(values, "A1")
    return sh.url

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Привет! Отправь PDF спецификацию. Я распознаю позиции и запишу их в Google Таблицу Import."
    )

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc:
        return
    if not doc.file_name.lower().endswith(".pdf"):
        await update.message.reply_text("Пока принимаю только PDF.")
        return

    await update.message.reply_text("Получил PDF. Распознаю и записываю в Google Sheets...")

    file = await doc.get_file()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_path = WORK_DIR / f"spec_{ts}.pdf"
    await file.download_to_drive(str(pdf_path))

    try:
        products, search, synonyms = load_database(TEMPLATE_FILE)
        text = extract_pdf_text(pdf_path)
        positions, review, unknown = recognize(text, products, search, synonyms)
        rows = add_default_works(positions + review + unknown, products)
        sheet_url = write_import_to_google_sheet(rows)

        await update.message.reply_text(
            f"Готово.\n"
            f"✅ Найдено автоматически: {len(positions)}\n"
            f"⚠️ На проверку: {len(review)}\n"
            f"❌ Не найдено: {len(unknown)}\n\n"
            f"Данные записаны в Google Таблицу Import:\n{sheet_url}"
        )
    except Exception as e:
        await update.message.reply_text(f"Ошибка: {e}")

def main():
    if not BOT_TOKEN:
        raise RuntimeError("Не найден BOT_TOKEN.")
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.run_polling()

if __name__ == "__main__":
    main()
